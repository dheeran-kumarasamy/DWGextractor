#!/usr/bin/env python3
import argparse
import csv
import os
import subprocess
import sys
import tempfile


def read_csv(path):
    with open(path, "r", encoding="utf-8", newline="") as f:
        raw = f.read()
    # Backward-compat: older extractor builds wrote literal "\\n" markers.
    if "\\n" in raw:
        raw = raw.replace("\\n", "\n")
    return list(csv.reader(raw.splitlines()))


def write_sheet(ws, rows):
    for row in rows:
        ws.append(row)


def main():
    parser = argparse.ArgumentParser(
        description="Export DWG entities and layers to an Excel workbook (.xlsx)."
    )
    parser.add_argument("input_dwg", help="Path to input DWG file")
    parser.add_argument("output_xlsx", help="Path to output XLSX file")
    parser.add_argument(
        "--extractor",
        default=os.path.join(os.path.dirname(__file__), "dwg_entity_types"),
        help="Path to dwg_entity_types binary",
    )
    args = parser.parse_args()

    if not os.path.isfile(args.input_dwg):
        print(f"Input DWG not found: {args.input_dwg}", file=sys.stderr)
        return 1

    if not os.path.isfile(args.extractor):
        print(
            f"Extractor binary not found: {args.extractor}\n"
            "Build it first (example):\n"
            "clang++ -std=c++14 -O2 -I src -I src/intern "
            "src/intern/*.cpp src/*.cpp tools/dwg_entity_types.cpp -o tools/dwg_entity_types",
            file=sys.stderr,
        )
        return 2

    try:
        from openpyxl import Workbook
    except Exception:
        print(
            "openpyxl is required. Install with: python3 -m pip install --user openpyxl",
            file=sys.stderr,
        )
        return 3

    with tempfile.TemporaryDirectory() as td:
        entities_csv = os.path.join(td, "entities.csv")
        layers_csv = os.path.join(td, "layers.csv")
        types_csv = os.path.join(td, "types.csv")

        cmd = [
            args.extractor,
            args.input_dwg,
            "--entities-csv",
            entities_csv,
            "--layers-csv",
            layers_csv,
            "--types-csv",
            types_csv,
        ]
        run = subprocess.run(cmd, capture_output=True, text=True)
        if run.returncode != 0:
            print("Extractor failed:", file=sys.stderr)
            if run.stdout:
                print(run.stdout, file=sys.stderr)
            if run.stderr:
                print(run.stderr, file=sys.stderr)
            return 4

        entities_rows = read_csv(entities_csv)
        layers_rows = read_csv(layers_csv)
        types_rows = read_csv(types_csv)

        wb = Workbook()
        ws_entities = wb.active
        ws_entities.title = "entities"
        write_sheet(ws_entities, entities_rows)

        ws_layers = wb.create_sheet("layers")
        write_sheet(ws_layers, layers_rows)

        ws_types = wb.create_sheet("entity_types")
        write_sheet(ws_types, types_rows)

        wb.save(args.output_xlsx)

    print(f"Excel written: {args.output_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
